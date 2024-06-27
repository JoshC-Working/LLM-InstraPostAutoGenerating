"""
At the command line, only need to run once to install the package via pip:

$ pip install google-generativeai
"""

import google.generativeai as genai
import pandas as pd
import time 
import openpyxl

from bs4 import BeautifulSoup
from markdown import markdown
from tool import copy_file
import re
from openpyxl.styles import Font
import os 


# def remove_markdown(text):
#     html = markdown.markdown(text)
#     plain_text = ''.join(plain_text for plain_text in html.itertext())
#     return plain_text

genai.configure(api_key="AIzaSyD2gXw17sww_XhHCmDFdxbyUAaYnsrDYtA")

# Set up the model
generation_config = {
	"temperature": 1,
	"top_p": 0.95,
	"top_k": 0,
	"max_output_tokens": 1000,
}

safety_settings = [
	{
		"category": "HARM_CATEGORY_HARASSMENT",
		"threshold": "BLOCK_NONE"
	},
	{
		"category": "HARM_CATEGORY_HATE_SPEECH",
		"threshold": "BLOCK_NONE"
	},
	{
		"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
		"threshold": "BLOCK_NONE"
	},
	{
		"category": "HARM_CATEGORY_DANGEROUS_CONTENT",
		"threshold": "BLOCK_NONE"
	},
]
# safety_settings = [
#   {
#     "category": "HARM_CATEGORY_HARASSMENT",
#     "threshold": "BLOCK_MEDIUM_AND_ABOVE"
#   },
#   {
#     "category": "HARM_CATEGORY_HATE_SPEECH",
#     "threshold": "BLOCK_MEDIUM_AND_ABOVE"
#   },
#   {
#     "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
#     "threshold": "BLOCK_MEDIUM_AND_ABOVE"
#   },
#   {
#     "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
#     "threshold": "BLOCK_MEDIUM_AND_ABOVE"
#   },
# ]

model = genai.GenerativeModel(model_name="gemini-1.5-pro-latest",
                        generation_config=generation_config,
                        safety_settings=safety_settings)           

chat = model.start_chat(history=[])


## Read Excel files
def read_excel_rows(file_path: str, has_header: bool = True) -> list:
	if has_header:  
		df = pd.read_excel(file_path)
	else:
		df = pd.read_excel(file_path, header = None )

	rows_list = df.values.tolist()
	return rows_list

## Write the result in 

def write_excel_to_rows(file_path, data_list, col_index, has_header= True):
	if has_header:
		offset = 2
	else:
		offset =1 
	# Load the Excel file

	workbook = openpyxl.load_workbook(file_path)

	# Select the first sheet
	sheet = workbook.active

	# Iterate through each row and write the list
	for row_index, row_data in enumerate(data_list):
		cell = sheet.cell(row=row_index + offset, column=col_index)
		cell.value = str(row_data)

	# Save the modified file
	workbook.save(file_path)


## 
def markdown_format_off(markdown_format_string):

	html = markdown(markdown_format_string)
	text = "".join(BeautifulSoup(html, features="html.parser").findAll(text=True))
	return text

##
def cross_check(topic:str, response:str, check_prompt1 ,check_prompt2,check_prompt3 ) -> int: 
  flag = False
  while (flag is False):
    time.sleep(time_stop)
    eval = model.generate_content(check_prompt1+topic+check_prompt2+str(response)+check_prompt3)
    time.sleep(time_stop)
    value = markdown_format_off(eval.text)
    print(value)
    if(len(value)<=3 ):
      flag = True

  print("Check_percentage: "+ eval.text)
  if int(value) >= 70:
     return 1
  else: 
     return 0

def save_2D_to_excel(arr_2d:list[list[str]], excel_path, offset_row , offset_col,font_size):

			
	wb = openpyxl.load_workbook(excel_path)   
	ws = wb.active
	font = Font(size=font_size) # font size
	for i, arr_1d in enumerate(arr_2d):
		for j, item in enumerate(arr_1d) :

			
			cell = ws.cell(row=offset_row+i, column= offset_col+j, value= arr_2d[i][j]) # Put the data into the cell
			cell.font = font
			time.sleep(1)
	wb.save(excel_path)


def split_response(text):
# Initialize the dictionary to store extracted information
    info = {"caption": "", "Main Points1": "", "Main Points2": "", "Main Points3": ""}

    # Define the regex patterns for capturing the required sections
    caption_pattern = r"[#*][#*] [cC]aption:(.*?)[#*][#*] [mM]ain [pP]oints:"
    # caption_pattern = r"## Caption:(.*?)## Main [pP]oints:"
    main_points_pattern = r"[#*][#*] [mM]ain [pP]oints:\s*1\.\s*(.*?)\s*2\.\s*(.*?)\s*3\.\s*(.*)"

    # Search for the patterns in the text
    caption_match = re.search(caption_pattern, text, re.DOTALL)
    main_points_match = re.search(main_points_pattern, text, re.DOTALL)

    # Check if caption is found
    if caption_match:
        info["caption"] = caption_match.group(1).strip()
    else:
        return None
    
    # Check if main points are found
    if main_points_match:
        info["Main Points1"] = "1. "+ main_points_match.group(1).strip()
        info["Main Points2"] = "2. "+ main_points_match.group(2).strip()
        info["Main Points3"] = "3. " + main_points_match.group(3).strip()
    else:
        return None

    # Return the information as a list
    return [info["caption"], info["Main Points1"], info["Main Points2"], info["Main Points3"]]

def save_to_excel(string_data, file_path, row, column, font_size: float):
    wb = openpyxl.load_workbook(file_path)    
    ws = wb.active

    font = Font(size=font_size) # font size

    cell = ws.cell(row=row, column=column, value=string_data) # Put the data into the cell
    cell.font = font
    
    wb.save(file_path)
    # print(f"Data saved to {file_path}")


contents = []

# comp_prompt_1 = '''Could you please write a description about the following topic?'''
# comp_prompt_2 = '''Only output the description.'''
comp_prompt_1 = ''''''
comp_prompt_2 = '''This is the title of a instagram post, could you generate a possible long captions(300 words) and 3 main points for the title? Like the format below: 
Caption: 
<Caption>
Main points:
1. <Point 1>
2. <Point 2>
3. <Point 3>
only output the caption and three bullet points'''
comp_prompt_3 = ""


titles = read_excel_rows("target.xlsx", has_header = True)
titles = [(i[0] if (len(i)>=1) else "") for i in titles]
# titles = ["How to launch a startup?", "Startup Funding Rounds", "Startup Funding Options"]
# titles = ["How to launch a startup?"]
target_file = "target.xlsx"
time_stop = 5





# comp_prompt = '''Give me an short introduction of the following country. Only incliude the introduction in the response.'''



print("Titles: ")
print(titles)
print("\n")


## Generatingq Response ## 
counter = 0 
check_values= [[],]

index = 0
while index < len(titles):

	counter += 1
	print("### Prompt: "+ str(index+1) +"_" +str(counter))


	## Prompt
	response = chat.send_message(comp_prompt_1+" "+ titles[index]+" "+comp_prompt_2+" "+comp_prompt_3)
	time.sleep(time_stop)

	## Response Information Checking
	print(response)
	print(response.prompt_feedback)
	print(response.candidates[0].finish_reason)
	print(response.candidates[0].safety_ratings)
	print("response valid: "+ str(len(response.parts))) # 1: non-empty response, 0: emtpy response


  ## Check if the return response has a text part
	if(len(response.parts)!=0):
		
		print(response.text)
		print("-"*20)
        
		split_content = split_response(response.text)

		if split_content != None:
			print("The response format: 1")
			
			for i in range(len(split_content)):    
				split_content[i] = markdown_format_off(split_content[i])
				if i != 0:
					split_content[i] = (str(i)+". "+split_content[i]).replace("\n", "")
			print("Type of split_content: "+ str(type(split_content[0])))
			contents.append(split_content)
				
			time.sleep(time_stop)
	
			index += 1 	
			counter = 0 
			comp_prompt_3 = ""
			
		else:
			print("The response format: 0")       
			# comp_prompt_3 = "Mind the format"
			chat.send_message("Please refere to the format strictly. Show the wordings: 'Caption' and 'Main Points' ")
			
			


			# print(response.text)
			# text = markdown_format_off(response.text)
			# print(type(text))
			# contents.append(text)
			# time.sleep(time_stop)
			# check_value = cross_check(titles[i],
			# 							text, 
			# 							'''The title: ''',
			# 							'''\nThe description: ''',
			# 							'''\nCould you give me a accuracy rate out of 100 of 
			# 							how correlated the description and the title are. Only output the the accuracy rate, which is in integer''' )
			# time.sleep(time_stop)
			# print("Check_value_"+str(i)+": "+str(check_value))
			# check_values[i].append(check_value) 

			# if (check_value) ==1:
			# 	i += 1
			# 	check_values.append([])
			# 	counter = 0

	print("index : "+str(index))


			

     
	# copy_file(target_file)
# write_excel_to_rows(target_file, , 2, has_header= True)


### Print the result ###
print("### Response ###")
for i, response in enumerate(contents):
	print(str(i) +". ")






### Store it into Excel ###
offset_col, offset_row = 1,1

save_2D_to_excel(contents,"new/target.xlsx",2 ,2, 20 )



