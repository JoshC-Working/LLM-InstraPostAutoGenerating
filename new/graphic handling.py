from PIL import Image, ImageDraw, ImageFont
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import pandas as pd
import textwrap
import os
from openpyxl import load_workbook
from PIL import Image as PILImage
import time

from openpyxl.drawing.image import Image as XLImage

def read_excel_rows(file_path, has_header = True):
    if has_header:  
      df = pd.read_excel(file_path)
    else:
      df = pd.read_excel(file_path, header = None )
    
    rows_list = df.values.tolist()
    return rows_list
import os
from PIL import Image



def resize_images_in_folder(folder_path, max_width, max_height):
    # 确保输出文件夹存在
    # output_folder = os.path.join(folder_path, "resized")
    # os.makedirs(output_folder, exist_ok=True)
    
    # 遍历文件夹中的所有 PNG 文件
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(".png"):
            image_path = os.path.join(folder_path, filename)
            with Image.open(image_path) as img:
                original_width, original_height = img.size
                
                # 计算比例
                ratio = min(max_width / original_width, max_height / original_height)
                new_width = int(original_width * ratio)
                new_height = int(original_height * ratio)
                
                # 调整图像大小
                resized_img = img.resize((new_width, new_height))
                
                # 保存调整大小后的图像到输出文件夹
                resized_image_path = folder_path+"/"+filename
                print(resized_image_path)
                resized_img.save(resized_image_path)
                print(f"Resized {filename} and saved to {resized_image_path}")


def insert_image_to_excel(image_path, excel_path, sheet_name, cell, sizing_percentage = None ):
    # 打開圖片
    image = Image.open(image_path)

    # 打開 Excel 文件
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # 將圖片插入到 Excel 中的指定單元格
    sheet[cell] = ""
    img = XLImage(image)
    img.anchor = cell

    if sizing_percentage is not None:
        img.width = int(img.width*sizing_percentage/100)
    
        img.height = int(img.height*sizing_percentage/100)

    sheet.add_image(img)

    # 保存 Excel 文件
    workbook.save(excel_path)
    print(f"圖片已成功插入到 {excel_path} 的 {sheet_name} 中的單元格 {cell}。")


def resize_image_proportionally(image_path, target_width, target_height):
    # 打开图像文件
    pil_image = PILImage.open(image_path)
    
    original_width, original_height = pil_image.size

    width_ratio = target_width / original_width
    height_ratio = target_height / original_height

    if width_ratio < height_ratio:
        resized_width = target_width
        resized_height = int(original_height * width_ratio)
    else:
        resized_height = target_height
        resized_width = int(original_width * height_ratio)
    pil_image = pil_image.resize((resized_width, resized_height), Image.LANCZOS)
    # pil_image.thumbnail((resized_width, resized_height))


    return pil_image

def insert_png_into_excel(folder_path, excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    
    for column in range(1, 10):  # 调整第1列到第3列的宽度
        ws.column_dimensions[chr(64 + column)].width = 80
# 设置多行的高度
    for row in range(2, 10):  # 调整第1行到第3行的高度
        ws.row_dimensions[row].height = 200
    
    col_offset = 6
    row_offset = 2

    # 获取文件夹中的所有PNG文件
    imgs = [f for f in os.listdir(folder_path) if f.endswith('.png')]
    print(imgs)
    for img in imgs:

        post_id = int(img.split('-')[0][-1])
        page_id = int(img.split('-')[1].split('.')[0])

        img_path = os.path.join(folder_path, img)
        
        pil_image = PILImage.open(img_path)
        # pil_image = resize_image_proportionally(img_path, 100, 100)
        xl_image = XLImage(pil_image)

        col = col_offset + page_id
        row = row_offset + post_id

        cell = ws.cell(column = col-1, row = row-1) # -1: the column index
        ws.add_image(xl_image, cell.coordinate)       

    wb.save(excel_path)  
import os

def clear_folder(folder_path):
    # 获取文件夹中的所有文件和子文件夹的名称
    file_list = os.listdir(folder_path)
    
    # 遍历文件夹中的所有文件和子文件夹
    for file_name in file_list:
        # 拼接文件的完整路径
        file_path = os.path.join(folder_path, file_name)
        
        # 如果是文件，则删除文件
        if os.path.isfile(file_path):
            os.remove(file_path)
        # 如果是子文件夹，则递归调用 clear_folder() 函数删除子文件夹中的文件
        elif os.path.isdir(file_path):
            clear_folder(file_path)



def put_text(pic_name, text, font_size, padding, font_colour, padding_colour = "white", output_name = None,  align = "centre" , font_path = None, ):
    pic = Image.open("new/photos/"+pic_name)
    width1, img_h = pic.size

    font_path = font_path or "/System/Library/Fonts/Supplemental/Arial Bold.ttf"

    output_name = output_name or pic_name # Default: output_name = pic_name
    
    font = ImageFont.truetype(font_path, font_size)
    draw = ImageDraw.Draw(pic)

    # Text Handling 
    lines = textwrap.wrap(text, width = 40)
    print(lines)
    print("."*40)

    n = len(lines)

    left_padding = 40

    heights = [draw.textbbox((0, 0), lines[i], font=font)[2:][1] + padding * 2 + 40 for i in range(n)] 
    widths = [draw.textbbox((0, 0), lines[i], font=font)[2:][0] + padding * 2 for i in range(n)]

    for i in range(n):
        x = left_padding
        y = (img_h - sum(heights)) / 2 + sum(heights[:i])

        rect_x0 = x - padding
        rect_y0 = y - padding
        rect_x1 = x + widths[i] + padding
        rect_y1 = y + heights[i] - 20 + padding

        draw.rectangle([rect_x0, rect_y0, rect_x1, rect_y1], fill= padding_colour)
        draw.text((x, y), lines[i], font=font, fill=font_colour)

    """
    for i in range(n):
        text_width, text_height = draw.textbbox((0, 0), text[i], font=font)[2:] 
        print(text_width, text_height)  
        # Calculate position for the text
        if align == "centre":
            x = (width1 - text_width) / 2
            y = (img_h - (n)*(text_height+ 2* padding)) / 2 + (text_height + 2* padding)*i
        elif align == "left":
            x = left_padding
            y = (img_h - text_height) /2
        elif align == "right":
            x = width1 - left_padding-text_width
            y = (img_h - text_height) /2

        # Define rectangle bounds (adjust padding as needed)
        
        rect_x0 = x - padding
        rect_y0 = y - padding
        rect_x1 = x + text_width + padding
        rect_y1 = y + text_height + padding

        # Draw rectangle
        draw.rectangle([rect_x0, rect_y0, rect_x1, rect_y1], fill= background_colour)

        # Draw text
        draw.text((x, y), text[i], font=font, fill=font_colour)
    """

    # Save the image
    print(output_name)
    # print(os.path.dirname())
    pic.show()
    pic.save("new/output_photos/"+ output_name +".png")
    
    return pic


### Clear the Existing Photos ###
folder_path = "new/output_photos"
clear_folder(folder_path)



### Reading Files ###
titles = read_excel_rows("new/target.xlsx", has_header = True) 
titles = [(i[2:5].copy() if (len(i)>=1) else "") for i in titles]


titles = [["Page1", "Page2", "Page3"]]
print(titles)

img_arr = []
for i in range(len(titles)):
    img_arr.append([])

    for j in range(len(titles[i])):
        if j == 0:
            img = put_text("bg_1.jpg", str(titles[i][j]), 45,40, "black", "white", "photo"+str(i+1)+"-"+str(j+1),"centre")  
             
        else: 
            img = put_text("bg_2.jpg", str(titles[i][j]), 45,0, "black", "white", "photo"+str(i+1)+"-"+str(j+1),"centre")   

        img_arr[i].append(img)





# 设置参数
folder_path = "main/output_photos"  # 替换为你的文件夹路径
max_width = 300  # 最大宽度
max_height = 300  # 最大高度

# 调用函数
resize_images_in_folder("new/output_photos", 250,250)
# time.sleep(2)
# width = len(titles[-1])
# length = len(titles)
# insert_png_into_excel("new/output_photos", "new/target.xlsx")