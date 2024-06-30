"""
At the command line, only need to run once to install the package via pip:

$ pip install google-generativeai
"""
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl import load_workbook
import google.generativeai as genai
import pandas as pd
import time 
import openpyxl
from bs4 import BeautifulSoup
from markdown import markdown
from tool import copy_file


# def remove_markdown(text):
#     html = markdown.markdown(text)
#     plain_text = ''.join(plain_text for plain_text in html.itertext())
#     return plain_text

genai.configure(api_key="AIzaSyD2gXw17sww_XhHCmDFdxbyUAaYnsrDYtA")

# Set up the model
generation_config = {
  "temperature": 1,
  "top_p": 0.95,
  "top_k": 0,
  "max_output_tokens": 600,
}

safety_settings = [
  {
    "category": "HARM_CATEGORY_HARASSMENT",
    "threshold": "BLOCK_NONE"
  },
  {
    "category": "HARM_CATEGORY_HATE_SPEECH",
    "threshold": "BLOCK_NONE"
  },
  {
    "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
    "threshold": "BLOCK_NONE"
  },
  {
    "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
    "threshold": "BLOCK_NONE"
  },
]

model = genai.GenerativeModel(model_name="gemini-1.5-pro-latest",
                              generation_config=generation_config,
                              safety_settings=safety_settings)           

convo = model.start_chat(history=[
])


## Read Excel files
def read_excel_rows(file_path, has_header = True):
    if has_header:  
      df = pd.read_excel(file_path)
    else:
      df = pd.read_excel(file_path, header = None )
    
    rows_list = df.values.tolist()
    return rows_list

## Write the result in 

def write_excel_to_rows(file_path, data_list, col_index, has_header= True):
    if has_header:
       offset = 2
    else:
       offset = 1 
    # Load the Excel file

    workbook = openpyxl.load_workbook(file_path)

    # Select the first sheet
    sheet = workbook.active

    # Iterate through each row and write the list
    for row_index, row_data in enumerate(data_list):
        cell = sheet.cell(row=row_index + offset, column=col_index)
        cell.value = str(row_data)

    # Save the modified file
    workbook.save(file_path)

## Clean Blank Data
def clean_blank(arr):
    for i in range(len(arr)):
        for j in range(len(arr)):
            pass

## 
def markdown_format_off(markdown_format_string):

  html = markdown(markdown_format_string)
  text = "".join(BeautifulSoup(html, features="html.parser").findAll(text=True))
  return text

##
def cross_check(topic:str, response:str, check_prompt1 ,check_prompt2,check_prompt3 ) -> int: 
  flag = False
  while (flag is False):
    time.sleep(time_stop)
    eval = model.generate_content(check_prompt1+topic+check_prompt2+str(response)+check_prompt3)
    time.sleep(time_stop)
    value = markdown_format_off(eval.text)
    print(value)
    if(len(value)<=3 ):
      flag = True

  print("Check_percentage: "+ eval.text)
  if int(value) >= 70:
     return 1
  else: 
     return 0

##


def save_to_excel(string_data, file_path, row, column):
   
    wb = load_workbook(file_path)    

    ws = wb.active
    font = Font(size=18)

    cell = ws.cell(row=row, column=column, value=string_data)
    cell.font = font
    
 
    wb.save(file_path)
    print(f"Data saved to {file_path}")



##
def split_string(string):
   result = string.split("\n")
   return result

def clean_2D_arr(arr: list[list[str]]):
   for i, arr1 in enumerate(arr):
      for j, item in enumerate(arr1) :
            if item == "" or item == None:
                del arr[i][j]
            else:
               arr[i][j] = item.rstrip()
    


## Hyperparameter Initialization ## 

responses = []

# Based on the below context, extract exactly 3 very short bullet points. Do not give any explanation.
# Context:
# ````
# {context}
# ````

comp_prompt_1 = '''Based on the below context, extract exactly 3 very very short bullet points. Do not give any explanation. Context:'''
comp_prompt_2 = ''''''


target_file = "test.xlsx"
time_stop = 5



### Start ###
print("### Start ###")
titles = ["Navigating the exhilarating journey of launching a startup involves a potent blend of vision, planning, and resilience. It's about transforming a groundbreaking idea into a tangible reality, demanding meticulous market research, defining a unique value proposition, and crafting a comprehensive business plan. Securing funding through bootstrapping, angel investors, or venture capitalists is crucial for fueling growth. Building a passionate team, developing a minimum viable product, and marketing it effectively are key milestones. Embracing agility and adaptability amidst inevitable challenges, along with a relentless drive to learn and iterate, ultimately paves the way for startup success. "]
titles = read_excel_rows("main/target.xlsx", has_header = True) 
titles = [(i[1] if (len(i)>=1) else "") for i in titles] # Take the second column


print("### Titles")
print(titles)
print("\n\n")


### Generatingq Response ###
counter = 0
check_values= [[],]
i = 0
while i < len(titles):
  counter += 1
  print("### Prompt: "+ str(i+1) +"_" +str(counter))
  

  ## Prompt
  response = model.generate_content(comp_prompt_1+" "+ titles[i]+" "+comp_prompt_2)
  time.sleep(time_stop)

  ## Response Information Checking
  print(response)
  print(response.prompt_feedback)
  print(response.candidates[0].finish_reason)
  print(response.candidates[0].safety_ratings)
  print("response valid: "+ str(len(response.parts))) # 1: non-empty response, 0: emtpy response


  ## Check if the return response has a text part
  if(len(response.parts)!=0):
    text = markdown_format_off(response.text)
    print(type(text))
    responses.append(text)
    time.sleep(time_stop)
    i += 1
    counter = 0


  
  print("--"*30)
    
     
# copy_file(target_file)
# write_excel_to_rows("copy_"+ target_file, responses, 2, has_header= True)

for i in responses:
   print(i)
   print("--"*30)

split_responses = []

for i in responses:
    split_responses.append(split_string(i))


### Print the responses ###
print("### Print the responses ###")
print("Responses: "+str(split_responses))
clean_2D_arr(split_responses)
print("Responses aftering cleaning: "+str(split_responses))


### Store the responses ###

offset_x = 3 
offset_y = 2

for i, arr1 in enumerate(split_responses):
    for j, item in enumerate(arr1) :
   
        save_to_excel(split_responses[i][j],"main/target.xlsx" , offset_y+i, offset_x+j )

  
                     
   






















