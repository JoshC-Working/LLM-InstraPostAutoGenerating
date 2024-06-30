from PIL import Image, ImageDraw, ImageFont
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import pandas as pd
import textwrap
import os
from openpyxl import load_workbook
from PIL import Image as PILImage
import time

from openpyxl.drawing.image import Image as XLImage

def read_excel_rows(file_path, has_header = True):
    if has_header:  
      df = pd.read_excel(file_path)
    else:
      df = pd.read_excel(file_path, header = None )
    
    rows_list = df.values.tolist()
    return rows_list

def insert_image_to_excel(image_path, excel_path, sheet_name, cell, sizing_percentage = None ):
    # 打開圖片
    image = Image.open(image_path)

    # 打開 Excel 文件
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # 將圖片插入到 Excel 中的指定單元格
    sheet[cell] = ""
    img = XLImage(image)
    img.anchor = cell

    if sizing_percentage is not None:
        img.width = int(img.width*sizing_percentage/100)
    
        img.height = int(img.height*sizing_percentage/100)

    sheet.add_image(img)

    # 保存 Excel 文件
    workbook.save(excel_path)
    print(f"圖片已成功插入到 {excel_path} 的 {sheet_name} 中的單元格 {cell}。")


def resize_image_proportionally(image_path, target_width, target_height):
 
    pil_image = PILImage.open(image_path)
    return pil_image
    original_width, original_height = pil_image.size

    width_ratio = target_width / original_width
    height_ratio = target_height / original_height

    if width_ratio < height_ratio:
        resized_width = target_width
        resized_height = int(original_height * width_ratio)
    else:
        resized_height = target_height
        resized_width = int(original_width * height_ratio)

    pil_image.thumbnail((resized_width, resized_height))

    return pil_image

def insert_png_into_excel(folder_path, excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    
    for column in range(1, 10):  
        ws.column_dimensions[chr(64 + column)].width = 100

    for row in range(2, 10):  
        ws.row_dimensions[row].height = 500
    
    col_offset = 6
    row_offset = 2

    imgs = [f for f in os.listdir(folder_path) if f.endswith('.png')]
    print(imgs)
    for img in imgs:

        post_id = int(img.split('-')[0][-1])
        page_id = int(img.split('-')[1].split('.')[0])

        img_path = os.path.join(folder_path, img)
        
        # pil_image = PILImage.open(img_path)
        pil_image = resize_image_proportionally(img_path, 100, 100)
        xl_image = XLImage(pil_image)

        col = col_offset + page_id
        row = row_offset + post_id

        cell = ws.cell(column = col-1, row = row-1) # -1: the column index
        ws.add_image(xl_image, cell.coordinate)       

    wb.save(excel_path)  
import os

def clear_folder(folder_path):

    file_list = os.listdir(folder_path)
    

    for file_name in file_list:

        file_path = os.path.join(folder_path, file_name)
        
        if os.path.isfile(file_path):
            os.remove(file_path)
        
        elif os.path.isdir(file_path):
            clear_folder(file_path)



def put_text(name1, text, font_size, padding, font_colour, background_colour, image_name,  align = "centre" , font_path = None, ):
    img1 = Image.open("main/photos/"+name1)
    
    
    
    width1, img_h = img1.size

    font_path = font_path or "/System/Library/Fonts/Supplemental/Arial Bold.ttf"
    
    font = ImageFont.truetype(font_path, font_size)
    draw = ImageDraw.Draw(img1)

    # Text Handling 
    
    lines = textwrap.wrap(text, width = 40)
    print(lines)
    print("."*40)

    n = len(lines)
    # Get text size

    left_padding = 40

    heights = [draw.textbbox((0, 0), lines[i], font=font)[2:][1] + padding * 2 + 40 for i in range(n)]
    widths = [draw.textbbox((0, 0), lines[i], font=font)[2:][0] + padding * 2 for i in range(n)]

    for i in range(n):
        x = left_padding
        y = (img_h - sum(heights)) / 2 + sum(heights[:i])

        rect_x0 = x - padding
        rect_y0 = y - padding
        rect_x1 = x + widths[i] + padding
        rect_y1 = y + heights[i] - 20 + padding

        draw.rectangle([rect_x0, rect_y0, rect_x1, rect_y1], fill= background_colour)
        draw.text((x, y), lines[i], font=font, fill=font_colour)

    """
    for i in range(n):
        text_width, text_height = draw.textbbox((0, 0), text[i], font=font)[2:] 
        print(text_width, text_height)  
        # Calculate position for the text
        if align == "centre":
            x = (width1 - text_width) / 2
            y = (img_h - (n)*(text_height+ 2* padding)) / 2 + (text_height + 2* padding)*i
        elif align == "left":
            x = left_padding
            y = (img_h - text_height) /2
        elif align == "right":
            x = width1 - left_padding-text_width
            y = (img_h - text_height) /2

        # Define rectangle bounds (adjust padding as needed)
        
        rect_x0 = x - padding
        rect_y0 = y - padding
        rect_x1 = x + text_width + padding
        rect_y1 = y + text_height + padding

        # Draw rectangle
        draw.rectangle([rect_x0, rect_y0, rect_x1, rect_y1], fill= background_colour)

        # Draw text
        draw.text((x, y), text[i], font=font, fill=font_colour)
    """

    # Save the image
    print(image_name)
    # print(os.path.dirname())
    img1.save("main/output_photos/"+ image_name +".png")
    img1.show()
    return img1


output_path = "main/output_photos"

folder_path = "main/output_photos"
clear_folder(folder_path)

titles = read_excel_rows("main/target.xlsx", has_header = True) 
titles = [(i[2:5].copy() if (len(i)>=1) else "") for i in titles]
# titles = [["Page1", "Page2", "Page3"]]
print(titles)

img_arr = []
for i in range(len(titles)):
    img_arr.append([])

    for j in range(len(titles[i])):
        if j == 0:
            img = put_text("bg_1.jpg", str(j+1)+". "+str(titles[i][j]), 40,0, "black", "white", "photo"+str(i+1)+"-"+str(j+1),"centre")  
             
        else: 
            img = put_text("bg_2.jpg", str(j+1)+". "+str(titles[i][j]), 40,0, "black", "white", "photo"+str(i+1)+"-"+str(j+1),"centre")   

        img_arr[i].append(img)





if __name__ == "__main__":
    time.sleep(2)
    width = len(titles[-1])
    length = len(titles)
    insert_png_into_excel("main/output_photos", "main/target.xlsx")