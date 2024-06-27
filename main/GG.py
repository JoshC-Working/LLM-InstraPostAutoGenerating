from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import os

def insert_png_into_excel(folder_path, excel_path):
    # 加载现有的工作簿
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 获取文件夹中的所有PNG文件
    png_files = [f for f in os.listdir(folder_path) if f.endswith('.png')]
    
    # 初始化行和列索引
    row_index = 1
    column_index = 1
    
    # 遍历PNG文件并将它们插入到Excel表格的单元格中
    for png_file in png_files:
        # 拼接PNG文件的完整路径
        png_path = os.path.join(folder_path, png_file)
        
        # 从文件加载图像
        pil_image = PILImage.open(png_path)
        
        # 将PIL图像转换为openpyxl图像
        xl_image = XLImage(pil_image)
        
        # 添加图像到指定单元格
        cell = ws.cell(row=row_index, column=column_index)
        ws.add_image(xl_image, cell.coordinate)
        
        # 调整列索引，以便将下一个图像放置在下一列
        column_index += 1
        

    
    # 保存工作簿
    wb.save(excel_path)

# 调用函数，将指定文件夹中的PNG文件插入到现有的Excel表格中
folder_path = "main/output_photos"
excel_path = "main/target.xlsx"
insert_png_into_excel(folder_path, excel_path)
