import pandas as pd
import pandas as pd

def write_excel_rows(file_path, arr, col_no):
        # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)

    # Select the first sheet
    sheet = workbook.active

    # Iterate through each row and write the list
    for row_index, row_data in enumerate(arr):
        cell = sheet.cell(row=row_index + 1, column=2)
        cell.value = str(row_data)

    # Save the modified file
    workbook.save(file_path)

    # df = pd.read_excel(file_path)
    
    # num_columns = len(df.columns)
    # if col_no >= num_columns:
    #     num_columns_to_add = col_no - num_columns + 1
    #     for i in range(num_columns_to_add):
    #         df.insert(num_columns + i, f'Column {num_columns + i + 1}', '')
    
    # df.iloc[:, col_no] = arr
    # df.to_excel(file_path, index=False)

import openpyxl

def write_list_to_excel(file_path, data_list, col_index):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)

    # Select the first sheet
    sheet = workbook.active

    # Iterate through each row and write the list
    for row_index, row_data in enumerate(data_list):
        cell = sheet.cell(row=row_index + 1, column=col_index)
        cell.value = str(row_data)

    # Save the modified file
    workbook.save(file_path)
import shutil

def copy_file(source_path, destination_path):
    shutil.copy2(source_path, destination_path)


copy_file("target.xlsx", "target_return.xlsx")
write_excel_rows("target_return.xlsx", ["4", "2", "3"], 2)